from flask import Flask, app, render_template, request, redirect, url_for, session,jsonify 
import openpyxl
import json
 
app=Flask(__name__,static_url_path='/static')
 
# Load Excel data into session on app startup
def load_excel_data(): 
    excel_file = 'blackbox.xlsx'  # Provide the actual file path
    workbook = openpyxl.load_workbook(excel_file)
    with open('test.txt', 'w') as file:
        # Write a single line of text to the file
        file.write('"Introduction":{"answer": "wait a min", "image_paths": []}')
    excel_data = {}
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = {}
        for row in sheet.iter_rows(min_row=1, values_only=True):
            print("ROW content - ",row)
            is_yes_or_no, question, value, image_paths = row   
            if is_yes_or_no == "Yes":
                if not image_paths:
                    image_path_list = ['noimage.png']
                elif ';' in image_paths:
                    image_path_list = image_paths.split(';')
                else:
                    image_path_list = [image_paths]

                value = value.replace("'", "").replace('"', '').replace(':', '-')
                data[question] = {'value': value, 'image_paths': image_path_list}
        excel_data[sheet_name] = data
    #print("total data>>>>>>:", excel_data)
    return excel_data

# Load Excel data at app startup (just once)
excel_data = load_excel_data()
 
@app.route('/candidate')
def candidate():  
    submitted_data_dict = {}  # Define an empty dictionary by default
  
    with open('test.txt', 'r') as file: 
        submitted_data_str = file.read().replace("'","\"");
        #print(f"Data uuuuuu>>>>>>>>>>>>>>>>>>: {submitted_data_str}")
        try:
            submitted_data_dict = json.loads(submitted_data_str)
            # Now, 'submitted_data_dict' is a Python dictionary
            print(submitted_data_dict)
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")

         # Check the data type of submitted_data
        data_type = type(submitted_data_dict)
     

        # Print the data type to the server logs
        #print(f"Data Type>>>>>>>>>>>>>>>>>>: {data_type}")
        #print(f"Data>>>>>>>>>>>>>>>>>>: {submitted_data_dict}")
    return render_template('candidate.html', submitted_data=submitted_data_dict)

@app.route('/admin', methods=['GET', 'POST'])
def admin(): 
    if request.method == 'POST':
        
        print("Received data  >>>>>>  :", request.get_data())
 
        data = request.get_json()
        sheetname = data['sheetname']
        question = data['question']

        if sheetname in excel_data and question in excel_data[sheetname]:
            entry_data = excel_data[sheetname][question]
            answer = entry_data['value']
            image_paths = entry_data['image_paths']
            
        else:
            answer = None
            image_paths = None
      
        response_data = {'answer': answer, 'image_paths': image_paths}
         
        with open('test.txt', 'w') as file:
            #file.write(f'"Sheet Name": "{sheetname}", "Question":"{question}"')
            file.write('{')
            file.write(f'"{question}": {response_data}')
            file.write('}')

    return render_template('admin.html', excel_data=excel_data);

@app.route('/check_update')
def check_update(): 
    current_data_hash = hash(str(get_submitted_data()))
    print(current_data_hash)
    print(session.get('current_data_hash', None))
    if current_data_hash != session.get('current_data_hash', None):
        session['current_data_hash'] = current_data_hash
        return jsonify({'updated': True})
    
    return jsonify({'updated': False})

if __name__ == '__main__':
    app.run(debug=True)
